# -*- coding: utf-8 -*-
"""로컬 실연동 리허설 (임시). 실제 uploader 로직 → localhost API → DB 실왕복.
   Excel COM 대신 openpyxl 로 마스터 구조를 읽고 더미데이터 주입(파일 미저장, 마스터 무손상).
   실행 전 stocks rank1~3 스냅샷 → 리허설 → 원상복구."""
import os, sys, json, datetime
import openpyxl, requests

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
ENV  = os.path.join(ROOT, "web", ".env.local")
MASTER = os.path.join(HERE, "STARSTOCK_MASTER.xlsm")
LOCAL_API = "http://localhost:3000/api/upload-stocks"

def load_env(path):
    d = {}
    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line: continue
            k, v = line.split("=", 1)
            d[k.strip()] = v.strip().strip('"')
    return d

env = load_env(ENV)
SUPA = env["NEXT_PUBLIC_SUPABASE_URL"].rstrip("/")
SKEY = env["SUPABASE_SERVICE_ROLE_KEY"]
H = {"apikey": SKEY, "Authorization": f"Bearer {SKEY}", "Content-Type": "application/json"}

# ── 1) 스냅샷 (rank 1,2,3) ─────────────────────────────
print("=== 1) 스냅샷 ===")
snap = requests.get(f"{SUPA}/rest/v1/stocks?rank=in.(1,2,3)&select=*", headers=H, timeout=15).json()
snap_ranks = {r["rank"] for r in snap}
print(f"기존 rank1~3 존재: {sorted(snap_ranks)} ({len(snap)}행 백업)")

# ── 2) 더미데이터 주입 (openpyxl, 파일 미저장) ─────────
wb = openpyxl.load_workbook(MASTER)
ws = wb["MASTER"]
dummy = [
    # code,   name,        cur,    open,  high,  low,   target, stop,  J상태,     K메모,      L전일고가
    ("005930","삼성전자",   72000, 70000, 72500, 69800, 80000, 68000, "매수적기","돌파확인",  70000),
    ("000660","SK하이닉스", 185000,181000,186000,180500,210000,175000,"관망유지","수급양호",  180000),
    ("035420","NAVER",      210000,206000,211000,205500,240000,200000,"매수적기","신고가",    205000),
]
for i, d in enumerate(dummy):
    r = 11 + i
    for c, val in zip(range(2, 13), d):  # B(2)~L(12)
        ws.cell(r, c).value = val

# ── 3) FakeExcel 로 실제 업로더 로직 구동 ──────────────
class FakeExcel:
    def __init__(self, ws): self.ws = ws
    def connect(self): return True
    def cell_val(self, row, col):
        try: return self.ws.cell(row, col).value
        except Exception: return None
    def set_cell(self, row, col, value):
        try: self.ws.cell(row, col).value = value
        except Exception: pass
    def set_status(self, text): print(f"   [엑셀상태셀] {text}")
    def get_interval_minutes(self): return 3

sys.path.insert(0, HERE)
from starstock_uploader import StarStockUploader

u = StarStockUploader(os.path.join(HERE, "config.json"))
u.cfg["api_url"] = LOCAL_API                 # 로컬로 강제
u.xl = FakeExcel(ws)
u.tracker.path = os.path.join(HERE, "_rehearsal_tracker.json")
u.tracker.reset_all()

print("\n=== 2) 돌파 감지 (2캔들 시뮬) ===")
u.detect_and_update()                        # 캔들1 → count=1
confirmed = u.detect_and_update()            # 캔들2 → count=2 → 확정
print(f"확정 종목 수: {len(confirmed)}  ranks={[s['rank'] for s in confirmed]}")

print("\n=== 3) 로컬 API 전송 ===")
ok = u.upload(confirmed)
print(f"업로드 결과: {'성공(200)' if ok else '실패'}")

# ── 4) DB 반영 검증 ────────────────────────────────────
print("\n=== 4) DB/상태 검증 ===")
st = requests.get("http://localhost:3000/api/stocks/status", timeout=15).json()
print("status API:", json.dumps(st, ensure_ascii=False))
rows = requests.get(f"{SUPA}/rest/v1/stocks?rank=in.(1,2,3)&select=rank,stock_code,stock_name,current_price,entry_confirmed,is_visible&order=rank", headers=H, timeout=15).json()
print("DB stocks rank1~3:", json.dumps(rows, ensure_ascii=False))

# ── 5) 원상복구 ────────────────────────────────────────
print("\n=== 5) 원상복구 ===")
# 리허설이 만든 rank 중 원래 없던 것 삭제
made = {1, 2, 3} - snap_ranks
for rk in made:
    requests.delete(f"{SUPA}/rest/v1/stocks?rank=eq.{rk}", headers=H, timeout=15)
# 원래 있던 행 되돌리기(upsert)
if snap:
    hh = dict(H); hh["Prefer"] = "resolution=merge-duplicates"
    requests.post(f"{SUPA}/rest/v1/stocks?on_conflict=rank", headers=hh, data=json.dumps(snap), timeout=15)
print(f"삭제된 rank(원래없음): {sorted(made)} / 복원된 rank(원래있음): {sorted(snap_ranks)}")
after = requests.get(f"{SUPA}/rest/v1/stocks?rank=in.(1,2,3)&select=rank,stock_code&order=rank", headers=H, timeout=15).json()
print("복구 후 rank1~3:", json.dumps(after, ensure_ascii=False))
print("\n리허설 완료. (마스터 엑셀/실 config 무손상, DB 원상복구)")
